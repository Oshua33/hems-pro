import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import requests
from utils import BASE_URL
import datetime
from datetime import datetime
from datetime import date

import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
#from customtkinter import CTkMessagebox
from customtkinter import CTkImage

from openpyxl.utils import get_column_letter


from tkinter import Tk, Button, messagebox
from utils import export_to_excel, print_excel
import requests
import os
import sys
import pandas as pd
from payment_gui import PaymentManagement  # Import the Payment GUI
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import tempfile
import os
import platform
import openpyxl
from PIL import Image, ImageTk
import io
import requests



from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import webbrowser
from tkinter import filedialog
import tkinter.simpledialog as simpledialog

import tempfile
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
import webbrowser
import os

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from tkinter import messagebox
from datetime import datetime






# Global hotel name
HOTEL_NAME = "Destone Hotel & Suite"

class RoundedButton(tk.Canvas):
    def __init__(self, parent, text, command, radius=12, padding=5, 
                 color="#34495E", hover_color="#1ABC9C", text_color="white", 
                 font=("Helvetica", 10), border_color="#16A085", border_width=2):
        self.command = command
        self.radius = radius
        self.color = color
        self.hover_color = hover_color
        self.text_color = text_color
        self.padding = padding
        self.font = font
        self.text = text
        self.border_color = border_color
        self.border_width = border_width

        self.current_view = "bookings"

        
        # Adjusted button size
        self.width = 120  # Smaller width
        self.height = 30  # Smaller height

        super().__init__(parent, width=self.width, height=self.height, bg=parent["bg"], highlightthickness=0)

        # Draw rounded rectangle and text
        self.rounded_rect = self.create_round_rect(5, 5, self.width-5, self.height-5, self.radius, fill=self.color, outline=self.border_color, width=self.border_width)
        self.text_id = self.create_text(self.width//2, self.height//2, text=self.text, fill=self.text_color, font=self.font)

        # Bind Events
        self.tag_bind(self.rounded_rect, "<Enter>", self.on_enter)
        self.tag_bind(self.rounded_rect, "<Leave>", self.on_leave)
        self.tag_bind(self.rounded_rect, "<Button-1>", self.on_click)
        self.tag_bind(self.text_id, "<Enter>", self.on_enter)
        self.tag_bind(self.text_id, "<Leave>", self.on_leave)
        self.tag_bind(self.text_id, "<Button-1>", self.on_click)

    def create_round_rect(self, x1, y1, x2, y2, r=25, **kwargs):
        points = [
            x1+r, y1,
            x2-r, y1,
            x2, y1,
            x2, y1+r,
            x2, y2-r,
            x2, y2,
            x2-r, y2,
            x1+r, y2,
            x1, y2,
            x1, y2-r,
            x1, y1+r,
            x1, y1
        ]
        return self.create_polygon(points, smooth=True, splinesteps=36, **kwargs)

    def on_enter(self, event=None):
        self.itemconfig(self.rounded_rect, fill=self.hover_color)

    def on_leave(self, event=None):
        self.itemconfig(self.rounded_rect, fill=self.color)

    def on_click(self, event=None):
        if self.command:
            self.command()




# =================== Main Booking Management ===================
class BookingManagement:
    def __init__(self, root, token):
        self.root = tk.Toplevel(root)
        self.root.title("HEMS-Booking Management")
        self.root.state("zoomed")
        self.root.configure(bg="#f0f0f0")
        


        
        self.token = token
        self.username = "current_user"

        self.tree = ttk.Treeview(self.root)  # Treeview (make sure it's defined)
        self.attachment_path_var = ctk.StringVar()  # or Tkinter.StringVar()


        self.entries = {}  # Add your entry widgets here
        self.guest_search_index = 0
        self.guest_search_results = []
        self.last_guest_name = None
        self.current_attachment = None

        
        # Set application icon
        icon_path = os.path.abspath("frontend/icon.ico")
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)

        # Set window size and position
        window_width = 1375
        window_height = 587
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x_coordinate = (screen_width // 2) - (window_width // 2)
        y_coordinate = (screen_height // 2) - (window_height // 2)
        self.root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")


    
        # ========== Main Layout ==========
        self.container = tk.Frame(self.root, bg="#AAB1B3", padx=10, pady=10)  # Light background for overall app
        self.container.pack(fill=tk.BOTH, expand=True)

        # Header Frame
        self.header_frame = tk.Frame(self.container, bg="#2C3E50", height=45)
        self.header_frame.pack(fill=tk.X, pady=(0, 5))  # More breathing room after header
        self.header_frame.pack_propagate(False)  # Prevent auto resizing based on child widgets

        # Title Label (left side of header)
        self.title_label = tk.Label(
            self.header_frame,
            text="📅 Booking Management",
            font=("Helvetica", 16, "bold"),
            fg="gold",
            bg="#2C3E50",
            anchor="w"
        )
        self.title_label.pack(side=tk.LEFT, padx=20)  # Padding left so it doesn't stick to edge

        # Action Frame (right side of header)
        self.action_frame = tk.Frame(self.header_frame, bg="#2C3E50")
        self.action_frame.pack(side=tk.RIGHT, padx=20, pady=10)

        # Add buttons to action frame (example: Export and Print buttons)
        self.export_button = RoundedButton(
            self.action_frame,
            text="📊Export to Excel",
            command=lambda: self.export_report(),
            radius=10,
            color="#95A5A6",
            hover_color="#16A085",
            text_color="white",
            border_color="#16A085",
            font=("Helvetica", 9, "bold"),
            border_width=2
        )
        self.export_button.pack(side=tk.LEFT, padx=5)

        self.print_button = RoundedButton(
            self.action_frame,
            text="🖨 Print Report",
            command=lambda: self.print_report(),
            radius=10,
            color="#95A5A6",
            hover_color="#16A085",
            text_color="white",
            border_color="#16A085",
            font=("Helvetica", 9, "bold"),
            border_width=2
        )
        self.print_button.pack(side=tk.LEFT, padx=5)

        self.booking_fields = [
        "ID", "Room Number", "Guest Name", "Gender", "Booking Cost", "Arrival Date", "Departure Date", "Status",
        "Number of Days", "Booking Type", "Phone Number", "Booking Date", "Payment Status",
        "Mode of Identification", "Identification Number", "Address",
        "Vehicle No", "Attachment", "Created By"
    ]



       # ========== Main Content Frame (Sidebar + Right Content) ==========
        self.main_frame = tk.Frame(self.container, bg="#AAB1B3")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Sidebar Frame
        self.left_frame = tk.Frame(self.main_frame, bg="#2C3E50", width=140)  # Set desired width
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        self.left_frame.pack_propagate(False)  # Prevent frame from resizing

        # Sidebar Label
        self.menu_label = tk.Label(self.left_frame, text="MENU", font=("Helvetica", 12, "bold"), 
                                fg="white", bg="#34495E", pady=5)
        self.menu_label.pack(fill=tk.X)

        # Right Content Frame
        self.right_frame = tk.Frame(self.main_frame, bg="#D6D8DA", relief="ridge", borderwidth=2)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Subheading Label
        self.subheading_label = tk.Label(self.right_frame, text="Select an option",
                                        font=("Helvetica", 14, "bold"), fg="#2C3E50", bg="#D6D8DA")
        self.subheading_label.pack(pady=10)

    

        # Sidebar Buttons
        buttons = [
            ("Create Booking", self.create_booking),
            ("List Booking", self.list_bookings),
            ("Sort By Status", self.list_bookings_by_status),
            ("Sort Guest Name", self.search_booking_by_guest_name),
            #("Sort by ID", self.search_booking_by_id),
            ("Sort By Room", self.search_booking_by_room),
            #("Update Booking", self.update_booking),
            ("Guest Checkout", self.guest_checkout),
            ("Cancel Booking", self.cancel_booking),
        ]

        for text, command in buttons:
            rb = RoundedButton(
                self.left_frame,
                text=text,
                command=lambda t=text, c=command: self.update_subheading(t, c),
                radius=12,
                color="#34495E",
                hover_color="#1ABC9C",
                border_color="#1ABC9C",  # or any color you want
                font=("Helvetica", 9)
            )
            rb.pack(anchor="w", padx=(8, 0), pady=2)


            
            # Dashboard Link
            self.dashboard_label = tk.Label(
            self.left_frame,
            text="⬅ Dashboard",
            cursor="hand2",
            font=("Helvetica", 10, "bold"),
            fg="white",
            bg="#1A5276",  # Deep Blue Background
            padx=10,
            pady=5,
            relief="solid",
            borderwidth=2
        )
        self.dashboard_label.pack(pady=15, padx=10, anchor="w", fill="x")

        # Change background color when hovering over
        self.dashboard_label.bind("<Enter>", lambda e: self.dashboard_label.config(bg="#154360"))  # Darker Blue on Hover
        self.dashboard_label.bind("<Leave>", lambda e: self.dashboard_label.config(bg="#1A5276"))  # Reset on Leave

        # Click event to open dashboard
        self.dashboard_label.bind("<Button-1>", lambda e: self.open_dashboard_window())


    def update_subheading(self, text, command):
        """Updates the subheading label and runs the selected command"""
        self.subheading_label.config(text=text)
        for widget in self.right_frame.winfo_children():
            widget.destroy()
        command()
    
    def open_dashboard_window(self):
        """Opens the dashboard window"""
        from dashboard import Dashboard
        Dashboard(self.root, self.username, self.token)
        self.root.destroy()



    def apply_grid_effect(self, tree=None):
        if tree is None:
            tree = self.tree  # Default to main tree if none is provided
        
        for i, item in enumerate(tree.get_children()):
            if i % 2 == 0:
                tree.item(item, tags=("evenrow",))
            else:
                tree.item(item, tags=("oddrow",))

        tree.tag_configure("evenrow", background="#e5f5e8")  # medium gray "#d9d9d9"
        tree.tag_configure("oddrow", background="white")


    def open_dashboard_window(self):
        from dashboard import Dashboard  # Import here to avoid circular import issues
        Dashboard(self.root, self.username, self.token)
        self.root.destroy()

    def update_subheading(self, text, command):
        if self.subheading_label.winfo_exists():
            self.subheading_label.config(text=text)
        for widget in self.right_frame.winfo_children():
            widget.destroy()
        command()
        


    def select_attachment(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            self.attachment_full_path = file_path
            self.attachment_label.configure(text=os.path.basename(file_path))  # Optional UI feedback
        else:
            self.attachment_full_path = None
            self.attachment_label.configure(text="No file selected")
        
             
        self.fetch_and_display_bookings()

       # Export and Print Buttons in Header Section
        def on_enter(e):
            e.widget.config(bg="#1ABC9C", fg="white")  # Background changes on hover

        def on_leave(e):
            e.widget.config(bg="#2C3E50", fg="white")  # Restore default background & text color

        self.export_button = tk.Label(self.header_frame, text="Export to Excel", 
                                    fg="white", bg="#2C3E50", font=("Helvetica", 9, "bold"), 
                                    cursor="hand2", padx=10, pady=5)
        self.export_button.pack(side=tk.RIGHT, padx=10, pady=5)
        self.export_button.bind("<Enter>", on_enter)
        self.export_button.bind("<Leave>", on_leave)
        self.export_button.bind("<Button-1>", lambda e: self.export_report())  # Click event

        self.print_button = tk.Label(self.header_frame, text="Print Report", 
                                    fg="white", bg="#2C3E50", font=("Helvetica", 9, "bold"), 
                                    cursor="hand2", padx=10, pady=5)
        self.print_button.pack(side=tk.RIGHT, padx=10, pady=5)
        self.print_button.bind("<Enter>", on_enter)
        self.print_button.bind("<Leave>", on_leave)
        self.print_button.bind("<Button-1>", lambda e: self.print_report())  # Click event


     



    def reset_booking_form(self):
        """Clears all input fields in the booking form."""
        if hasattr(self, "entries"):
            for key, entry in self.entries.items():
                if isinstance(entry, ttk.Combobox):
                    entry.set("")  # Clear combobox selection
                elif isinstance(entry, DateEntry):
                    entry.set_date(datetime.date.today())  # Reset date to today
                elif isinstance(entry, tk.Entry):
                    entry.delete(0, tk.END)  # Clear text entry
                else:
                    print(f"Unknown entry type for {key}")

    





    def fetch_and_display_bookings(self):
        """Fetch booking data from the API"""
        url = "http://127.0.0.1:8000/bookings/list"
        headers = {"Authorization": f"Bearer {self.token}"}

        try:
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                self.bookings_data = response.json()
            else:
                self.bookings_data = []
                messagebox.showerror("Error", "Failed to fetch bookings.")

        except Exception as e:
            self.bookings_data = []
            messagebox.showerror("Error", f"API Error: {str(e)}")

    

    def export_report(self):
        """Export current booking view to Excel with styled formatting, summary, and timestamped filename."""
        try:
            view_config = {
                "bookings": {
                    "tree": getattr(self, "tree", None),
                    "base_filename": "Booking_Report",
                    "total_cost": getattr(self, "total_booking_cost_label", None),
                    "total_entries": getattr(self, "total_entries_label", None)
                },
                "guest_search": {
                    "tree": getattr(self, "search_tree", None),
                    "base_filename": "Guest_Search_Report",
                    "total_cost": getattr(self, "total_cost_label", None),
                    "total_entries": getattr(self, "total_entries_label", None)
                },
                "room_search": {
                    "tree": getattr(self, "search_tree", None),
                    "base_filename": "Room_Search_Report",
                    "total_cost": getattr(self, "total_label", None),
                    "total_entries": getattr(self, "total_entries_label", None)
                },
                "status_search": {
                "tree": getattr(self, "tree", None),
                "base_filename": "Guest_Status_Search_Report",
                "total_cost": getattr(self, "total_cost_label", None),  # safer access
                "total_entries": getattr(self, "total_entries_label", None)
            }


            }

            config = view_config.get(self.current_view)
            if not config:
                messagebox.showerror("Export Error", f"No exportable view found for: {self.current_view}")
                return

            tree = config["tree"]
            if not tree or not tree.get_children():
                messagebox.showerror("Export Error", "No data available to export.")
                return

            base_filename = config["base_filename"]
            total_cost_text = config["total_cost"].cget("text") if config["total_cost"] else "Total Cost: N/A"
            total_entries_text = config["total_entries"].cget("text") if config["total_entries"] else "Total Entries: N/A"

            # Determine save path
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))

            project_root = os.path.abspath(os.path.join(script_dir, os.pardir))
            report_dir = os.path.join(project_root, "Reports")
            os.makedirs(report_dir, exist_ok=True)

            # Generate filename
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            filename = f"{base_filename}_{timestamp}.xlsx"
            file_path = os.path.join(report_dir, filename)

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reports"

            columns = [tree.heading(col)["text"] for col in tree["columns"]]

            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = "Hotel Booking Report"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")

            # Header
            ws.append(columns)
            for cell in ws[2]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )

            # Data
            for row_id in tree.get_children():
                values = tree.item(row_id)["values"]
                ws.append(values)
                row_idx = ws.max_row
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal="left")
                    cell.border = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin")
                    )

            # Autofit
            for col_idx, column in enumerate(columns, 1):
                max_len = len(str(column))
                for row_id in tree.get_children():
                    cell_val = str(tree.item(row_id)["values"][col_idx - 1])
                    if len(cell_val) > max_len:
                        max_len = len(cell_val)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max_len + 2

            # Summary
            summary_start_row = ws.max_row + 2
            summary_font = Font(bold=True, size=12)

            total_cost_cell = ws.cell(row=summary_start_row, column=1)
            total_cost_cell.value = total_cost_text
            total_cost_cell.font = summary_font

            total_entries_cell = ws.cell(row=summary_start_row + 1, column=1)
            total_entries_cell.value = total_entries_text
            total_entries_cell.font = summary_font

            # Save and open
            wb.save(file_path)
            messagebox.showinfo("Export Success", f"Report exported successfully to:\n{file_path}")
            os.startfile(file_path)

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Export Error", f"Failed to export report:\n{str(e)}")



    

    def print_report(self):
        """Print the exported Excel report"""
        if hasattr(self, 'last_exported_file') and self.last_exported_file:
            print_excel(self.last_exported_file)
        else:
            messagebox.showwarning("Warning", "Please export the report before printing.")

    def update_subheading(self, text, command):
        """Updates the subheading label and calls the selected function."""
        if hasattr(self, "subheading_label") and self.subheading_label.winfo_exists():
            self.subheading_label.config(text=text)
        else:
            self.subheading_label = tk.Label(self.right_frame, text=text, font=("Arial", 14, "bold"), bg="#f0f0f0")
            self.subheading_label.pack(pady=10)

        # Clear right frame before displaying new content
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        command()

    


    def create_booking(self, booking_data=None):
        """Opens a pop-up window for creating or updating a booking with CustomTkinter."""
        self.clear_right_frame()

        title_text = "Update Booking" if booking_data else "Create Booking"

        # Fields for both create and update modes
        fields = [
            ("Room Number", "room_number"),
            ("Guest Name", "guest_name"),
            ("Gender", "gender"),
            ("Booking Cost", "booking_cost"),
            ("Arrival Date", "arrival_date"),
            ("Departure Date", "departure_date"),
            ("Booking Type", "booking_type"),
            ("Phone Number", "phone_number"),
            ("Mode of Identification", "mode_of_identification"),
            ("Identification Number", "identification_number"),
            ("Address", "address"),
            ("Vehicle No", "vehicle_no"),
            ("Attachment", "attachment"),
        ]

        # Create a new pop-up window
        create_window = ctk.CTkToplevel(self.root)
        create_window.title(title_text)
        create_window.geometry("790x400")
        create_window.resizable(False, False)
        create_window.configure(fg_color="#f5f5f5")

        # Center the window on the screen
        screen_width = create_window.winfo_screenwidth()
        screen_height = create_window.winfo_screenheight()
        x_coordinate = (screen_width - 650) // 2
        y_coordinate = (screen_height - 400) // 2
        create_window.geometry(f"790x400+{x_coordinate}+{y_coordinate}")

        # Make the window modal
        create_window.transient(self.root)
        create_window.grab_set()

        # Header
        header_frame = ctk.CTkFrame(create_window, fg_color="#2c3e50", height=50, corner_radius=8)
        header_frame.pack(fill="x", padx=10, pady=10)
        header_label = ctk.CTkLabel(header_frame, text=title_text, font=("Arial", 16, "bold"), text_color="white")
        header_label.pack(pady=10)

        # Main Content Frame
        frame = ctk.CTkFrame(create_window, fg_color="#eeeeee", corner_radius=10)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.entries = {}  # Store entry widgets
        combo_box_values = {
            "Gender": ["Male", "Female"],
            "Booking Type": ["checked-in", "reservation", "complimentary"],
            "Mode of Identification": ["National Id Card", "Voter Card", "Id Card", "Passport"]
        }

        # Form Grid Layout
        form_data = [
            ("Room Number", ctk.CTkEntry, 0, 0),
            ("Guest Name", ctk.CTkEntry, 0, 2),
            ("Identification Number", ctk.CTkEntry, 1, 0),
            ("Mode of Identification", ctk.CTkComboBox, 1, 2),
            ("Gender", ctk.CTkComboBox, 2, 0),
            ("Address", ctk.CTkEntry, 2, 2),
            ("Phone Number", ctk.CTkEntry, 3, 0),
            ("Booking Type", ctk.CTkComboBox, 3, 2),
            ("Arrival Date", DateEntry, 4, 0),
            ("Departure Date", DateEntry, 4, 2),
            ("Vehicle No", ctk.CTkEntry, 5, 0),
            ("Attachment", ctk.CTkEntry, 5, 2),
        ]

        # Field-specific widths
        field_widths = {
            "Room Number": 50,
            "Guest Name": 170,
            "Identification Number": 150,
            "Mode of Identification": 130,
            "Gender": 100,
            "Address": 200,
            "Phone Number": 120,
            "Booking Type": 120,
            "Arrival Date": 10,
            "Departure Date": 10,
            "Vehicle No": 120,
            "Attachment": 200
        }

        # Form creation loop
        for label_text, field_type, row, col in form_data:
            label = ctk.CTkLabel(frame, text=label_text, font=("Helvetica", 12, "bold"), text_color="#2c3e50")
            label.grid(row=row, column=col, sticky="w", pady=5, padx=10)

            width = field_widths.get(label_text, 20)
            entry = None

            if field_type == ctk.CTkComboBox:
                entry = ctk.CTkComboBox(frame, values=combo_box_values.get(label_text, []), state="readonly", font=("Helvetica", 12), width=width)
            elif field_type == DateEntry:
                entry = DateEntry(frame, font=("Helvetica", 12), background='darkblue', foreground='white', borderwidth=2)
            else:
                entry = field_type(frame, font=("Helvetica", 12), width=width)

            # Handle Attachment field
            if label_text == "Attachment" and not booking_data:
                entry.insert(0, "Select file...")
                entry.configure(text_color="gray")

            entry.grid(row=row, column=col + 1, pady=5, padx=10, sticky="w")
            self.entries[label_text] = entry

            # Pre-fill entry if booking_data is provided
            if booking_data:
                key = label_text.lower().replace(" ", "_")
                value = booking_data.get(key, "")
                try:
                    entry.set(value)  # For CTkComboBox or DateEntry
                except:
                    entry.insert(0, value)  # For CTkEntry

        # Define file picker function and bind it
        def browse_file(event=None):
            file_path = filedialog.askopenfilename()
            if file_path:
                self.attachment_full_path = file_path  # Store full path for later upload
                attachment_entry.configure(state="normal")
                attachment_entry.delete(0, "end")
                attachment_entry.insert(0, file_path)

        attachment_entry = self.entries["Attachment"]
        attachment_entry.bind("<1>", browse_file)

        # Submit Button
        submit_btn = ctk.CTkButton(
            frame, text="Submit Booking", 
            command=lambda: self.submit_booking(create_window), 
            font=("Arial", 14, "bold"),
            fg_color="#3498db", hover_color="#2980b9", 
            text_color="white", corner_radius=10, 
            width=200, height=40
        )
        submit_btn.grid(row=7, column=0, columnspan=5, pady=25, sticky="n")

        # Search button for guest name
        guest_name_entry = self.entries["Guest Name"]

        # Set placeholder text and color for Guest Name
        guest_name_entry.insert(0, "Guest name")
        guest_name_entry.configure(text_color="gray")

        # Define focus-in and focus-out behaviors
        def clear_placeholder(event):
            if guest_name_entry.get() == "Guest name":
                guest_name_entry.delete(0, "end")
                guest_name_entry.configure(text_color="black")

        def add_placeholder(event):
            if guest_name_entry.get() == "":
                guest_name_entry.insert(0, "Guest name")
                guest_name_entry.configure(text_color="gray")

        guest_name_entry.bind("<FocusIn>", clear_placeholder)
        guest_name_entry.bind("<FocusOut>", add_placeholder)

        search_btn = ctk.CTkButton(
        frame,
        text="🔍Search",
        command=lambda: self.search_guest(guest_name_entry.get()),
        font=("Arial", 12),
        fg_color="gray",
        text_color="white",
        width=90,
        height=28
    )
        search_btn.grid(row=0, column=4, padx=(5, 0), pady=5, sticky="w")

       



    def search_guest(self, guest_name):
        if not hasattr(self, "entries"):
            messagebox.showerror("Error", "Form is not initialized properly.")
            return

        if not guest_name.strip():
            messagebox.showwarning("Input Error", "Please enter a guest name to search.")
            return

        token = getattr(self, "token", None)
        if not token:
            messagebox.showwarning("Authentication Error", "Please log in first.")
            return

        is_new_search = not hasattr(self, "last_guest_name") or self.last_guest_name != guest_name

        if is_new_search:
            try:
                headers = {"Authorization": f"Bearer {token}"}
                response = requests.get(
                    "http://localhost:8000/bookings/search-guest/",
                    params={"guest_name": guest_name},
                    headers=headers
                )
                response.raise_for_status()
                self.guest_search_results = response.json()
                self.guest_search_index = 0
                self.last_guest_name = guest_name

                if not self.guest_search_results:
                    messagebox.showinfo("Not Found", "Guest Name not found.")
                    return

                # Close existing popup if it exists
                if hasattr(self, "search_popup") and self.search_popup.winfo_exists():
                    self.search_popup.destroy()

                if len(self.guest_search_results) > 1:
                    self.show_guest_search_popup()

            except requests.HTTPError as e:
                if e.response.status_code == 404:
                    messagebox.showinfo("Not Found", "Guest Name not found.")
                else:
                    messagebox.showerror("Error", f"Failed to fetch guest info: {str(e)}")
                return
        else:
            self.guest_search_index = (self.guest_search_index + 1) % len(self.guest_search_results)

        selected_guest = self.guest_search_results[self.guest_search_index]

        # ✅ Fill form
        self.entries["Gender"].set(selected_guest.get("gender", ""))
        self.entries["Mode of Identification"].set(selected_guest.get("mode_of_identification", ""))
        self.entries["Identification Number"].delete(0, "end")
        self.entries["Identification Number"].insert(0, selected_guest.get("identification_number", ""))
        self.entries["Address"].delete(0, "end")
        self.entries["Address"].insert(0, selected_guest.get("address", ""))
        self.entries["Phone Number"].delete(0, "end")
        self.entries["Phone Number"].insert(0, selected_guest.get("phone_number", ""))
        self.entries["Vehicle No"].delete(0, "end")
        self.entries["Vehicle No"].insert(0, selected_guest.get("vehicle_no", ""))

        attachment_url = selected_guest.get("attachment", "")
        self.entries["Attachment"].delete(0, "end")
        self.entries["Attachment"].insert(0, attachment_url)

        self.current_attachment = attachment_url

        # ✅ Update popup count (if open)
        if hasattr(self, "search_popup") and self.search_popup.winfo_exists():
            self.search_popup_label.config(
                text=f"Showing result {self.guest_search_index + 1} of {len(self.guest_search_results)}"
            )

    def show_guest_search_popup(self):
        if hasattr(self, "search_popup") and self.search_popup.winfo_exists():
            # If popup already exists, just update the label text
            self.search_popup_label.config(
                text=f"Showing result {self.guest_search_index + 1} of {len(self.guest_search_results)}"
            )
            return

        self.search_popup = tk.Toplevel(self.root)
        self.search_popup.title("Search Info")

        # Set small fixed size and top-left position (e.g., x=10, y=10)
        self.search_popup.geometry("200x60+160+150")
        self.search_popup.resizable(False, False)
        self.search_popup.attributes('-topmost', True)

        self.search_popup_label = tk.Label(
            self.search_popup,
            text=f"Showing result {self.guest_search_index + 1} of {len(self.guest_search_results)}",
            font=("Arial", 9),
            padx=5, pady=5
        )
        self.search_popup_label.pack()

    # Optional: Auto-close after a few seconds
    # self.search_popup.after(5000, self.search_popup.destroy)


    def submit_booking(self, create_window, booking_data=None):
        try:
            is_update = booking_data is not None
            url = (
                f"http://127.0.0.1:8000/bookings/update/{booking_data['id']}/"
                if is_update else
                "http://127.0.0.1:8000/bookings/create/"
            )
            headers = {"Authorization": f"Bearer {self.token}"}

            # Fetch values
            attachment_text = self.entries["Attachment"].get().strip()
            attachment_path = getattr(self, "attachment_full_path", None)

            data = {
                "room_number": self.entries["Room Number"].get().strip().lower(),
                "guest_name": self.entries["Guest Name"].get(),
                "gender": self.entries["Gender"].get(),
                "mode_of_identification": self.entries["Mode of Identification"].get(),
                "identification_number": self.entries["Identification Number"].get(),
                "address": self.entries["Address"].get(),
                "arrival_date": self.entries["Arrival Date"].get_date().strftime("%Y-%m-%d"),
                "departure_date": self.entries["Departure Date"].get_date().strftime("%Y-%m-%d"),
                "booking_type": self.entries["Booking Type"].get(),
                "phone_number": self.entries["Phone Number"].get(),
                "vehicle_no": self.entries["Vehicle No"].get() or None,
            }

            # Validate required fields
            required_fields = ["room_number", "guest_name", "gender", "mode_of_identification",
                            "address", "arrival_date", "departure_date", "booking_type", "phone_number"]
            missing = [k for k in required_fields if not data.get(k)]
            if missing:
                messagebox.showerror("Missing Fields", f"The following fields are required:\n{', '.join(missing)}")
                return

            # --- Attachment Logic ---
            files = {}
            file_obj = None

            if attachment_path and os.path.isfile(attachment_path):
                # ✅ Upload file
                try:
                    file_obj = open(attachment_path, "rb")
                    files["attachment_file"] = (
                        os.path.basename(attachment_path),
                        file_obj,
                        "application/octet-stream"
                    )
                    # No need to set attachment_str if uploading file
                except Exception as e:
                    messagebox.showerror("File Error", f"Attachment file error: {str(e)}")
                    return
            elif attachment_text:
                # ✅ Send as string path (from fetched data or user input)
                data["attachment_str"] = attachment_text
            else:
                # ✅ Send None explicitly (user cleared it)
                data["attachment_str"] = None

            # Make request
            try:
                if is_update:
                    response = requests.put(url, data=data, files=files if files else None, headers=headers)
                else:
                    response = requests.post(url, data=data, files=files if files else None, headers=headers)
            finally:
                if file_obj:
                    file_obj.close()

            # Response Handling
            if response.status_code == 200:
                messagebox.showinfo("Success", f"Booking {'updated' if is_update else 'created'} successfully.")
                

                # ✅ Clear any cached file path after successful booking
                self.attachment_full_path = None

                # ✅ Also clear the Attachment entry if needed
                self.entries["Attachment"].delete(0, "end")
                create_window.destroy()

                # Close guest search popup if it's open
                if hasattr(self, "search_popup") and self.search_popup.winfo_exists():
                    self.search_popup.destroy()

            

            else:
                messagebox.showerror("Error", f"Failed to {'update' if is_update else 'create'} booking:\n{response.text}")

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Request Error", str(e))
        except Exception as e:
            messagebox.showerror("Unexpected Error", str(e))





    def list_bookings(self):
        self.current_view = "bookings"
        self.clear_right_frame()
        
        
        # Create a new frame for the table
        frame = tk.Frame(self.right_frame, bg="#D6D8DA", padx=5, pady=5)
        #frame = tk.Frame(self.right_frame, bg="#ffffff")
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="List Bookings Report", font=("Arial", 14, "bold"), bg="#D6D8DA").pack(pady=10)

        filter_frame = tk.Frame(frame, bg="#ffffff")
        filter_frame.pack(pady=5)

        tk.Label(filter_frame, text="Start Date:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=0, padx=5, pady=5)
        self.start_date = DateEntry(filter_frame, font=("Arial", 11))
        self.start_date.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(filter_frame, text="End Date:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=2, padx=5, pady=5)
        self.end_date = DateEntry(filter_frame, font=("Arial", 11))
        self.end_date.grid(row=0, column=3, padx=5, pady=5)

        fetch_btn = ttk.Button(
            filter_frame,
            text="Fetch Bookings",
            command=lambda: self.fetch_bookings(self.start_date, self.end_date)
        )
        fetch_btn.grid(row=0, column=4, padx=10, pady=5)
# === Table Frame ===
        table_frame = tk.Frame(frame, bg="#ffffff", bd=1, relief="solid")
        table_frame.pack(fill=tk.BOTH, expand=True, pady=1)

        # Create Treeview
        columns = ("ID", "Room No", "Guest Name", "Gender", "Booking Cost", "Arrival", "Departure", "Status", "Days", 
                "Booking Type", "Phone Number", "Booking Date", "Payment Status", "Mode of Identification", "Identification No", 
                "Address", "Vehicle No", "Attachment", "Created_by")

        style = ttk.Style()
        style.configure("Treeview", rowheight=25, background="white", fieldbackground="white", borderwidth=1)
        style.configure("Treeview.Heading", font=("Arial", 11, "bold"), background="#2c3e50", foreground="white")
        style.map("Treeview", background=[("selected", "#b3d1ff")])

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col, anchor="center")
            self.tree.column(col, width=70, anchor="center")  # Widened slightly for readability

        # Add scrollbars INSIDE same frame
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        # Pack them properly
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)


        # Label to display total booking cost
        self.total_booking_cost_label = tk.Label(frame, text="", font=("Arial", 12, "bold"), bg="#D6D8DA", fg="red")
        self.total_booking_cost_label.pack(pady=10)

        # Total Entries Label (Blue)
        self.total_entries_label = tk.Label(frame, text="Total Entries: 0", fg="blue", font=("Arial", 12, "bold"), bg="#D6D8DA")
        self.total_entries_label.pack(side="left", padx=(0, 10))


            
        # Assuming `frame` is where your booking buttons are being placed
        button_frame = ctk.CTkFrame(frame, fg_color="transparent")
        button_frame.pack(pady=10)

        view_button = ctk.CTkButton(button_frame, text="View Booking", corner_radius=20, command=self.view_selected_booking)
        view_button.pack(side="left", padx=10)


        update_button = ctk.CTkButton(button_frame, text="Update Booking", corner_radius=20, command=self.update_selected_booking)
        update_button.pack(side="left", padx=10)



    def update_selected_booking(self):
        selected_booking = self.get_selected_booking()
        if selected_booking:
            self.update_booking_form(selected_booking)

        else:
            messagebox.showwarning("No Selection", "Please select a booking to update.")

    def get_selected_booking(self):
        selected_item = self.tree.focus()
        if not selected_item:
            return None

        values = self.tree.item(selected_item, "values")

        field_keys = [
            "booking_id", "room_number", "guest_name", "gender", "booking_cost", "arrival_date",
            "departure_date", "status", "no_of_days", "booking_type", "phone_number", "booking_date",
            "payment_status", "mode_of_identification", "identification_number", "address",
            "vehicle_no", "attachment", "created_by"
        ]

        return dict(zip(field_keys, values))
    

    def update_booking_form(self, booking_data):
        def browse_attachment_file(event=None):
            file_path = filedialog.askopenfilename(
                title="Select Attachment",
                filetypes=[("All Files", "*.*"), ("Images", "*.png;*.jpg;*.jpeg"), ("PDFs", "*.pdf")]
            )
            if file_path:
                attachment_entry = self.update_entries.get("Attachment")
                if attachment_entry:
                    attachment_entry.delete(0, "end")
                    attachment_entry.insert(0, file_path)
                    attachment_entry.configure(text_color="black")  # Make text visible

        window = ctk.CTkToplevel(self.root)
        window.title("Update Booking")
        window.geometry("750x450+150+30")
        window.configure(fg_color="#e0e0e0")  # Light-medium gray background
        window.resizable(False, False)

        title = ctk.CTkLabel(window, text="Update Booking", font=("Segoe UI", 18, "bold"), text_color="#1e1e1e")

        title.pack(pady=(10, 5))

        form_frame = ctk.CTkFrame(window, fg_color="#f5f5f5", corner_radius=15)
        form_frame.pack(fill="both", expand=True, padx=20, pady=15)

        for i in range(4):
            form_frame.grid_columnconfigure(i, weight=1)

        combo_box_values = {
            "Gender": ["Male", "Female"],
            "Booking Type": ["checked-in", "reservation", "complimentary"],
            "Mode of Identification": ["National Id Card", "Voter Card", "Id Card", "Passport"]
        }

        self.update_entries = {}

        def add_form_row(row, label1, field1, label2=None, field2=None):
            col = 0
            for label_text, field_type in [(label1, field1), (label2, field2) if label2 else (None, None)]:
                if label_text:
                    label = ctk.CTkLabel(form_frame, text=label_text, font=("Segoe UI", 11, "bold"), text_color="#333333")

                    label.grid(row=row, column=col, sticky="w", padx=10, pady=6)

                    width = 180
                    if field_type == ctk.CTkComboBox:
                        entry = ctk.CTkComboBox(
                        form_frame,
                        values=combo_box_values.get(label_text, []),
                        state="readonly",
                        font=("Segoe UI", 11),
                        width=width,
                        fg_color="#f8f8f8",
                        text_color="#222222",
                        dropdown_fg_color="#f0f0f0",
                        dropdown_text_color="#222222"
                    )


                    elif field_type == DateEntry:
                        entry = DateEntry(
                        form_frame,
                        font=("Segoe UI", 10),
                        background='white',
                        foreground='black',
                        borderwidth=2
                    )

                    else:
                        entry = field_type(
                        form_frame,
                        font=("Segoe UI", 13),
                        width=width,
                        fg_color="#f8f8f8",  # Entry background
                        text_color="#222222"  # Entry text
                    )

                        if label_text == "Attachment":
                            entry.insert(0, "Select file...")
                            entry.configure(text_color="black")
                            entry.bind("<Button-1>", browse_attachment_file)
                        if label_text == "Booking ID":
                            entry.configure(state="readonly")

                    entry.grid(row=row, column=col + 1, sticky="w", padx=10, pady=6)
                    self.update_entries[label_text] = entry
                    col += 2

        # Form layout rows
        add_form_row(0, "Booking ID", ctk.CTkEntry)
        add_form_row(1, "Room Number", ctk.CTkEntry, "Guest Name", ctk.CTkEntry)
        add_form_row(2, "Identification Number", ctk.CTkEntry, "Mode of Identification", ctk.CTkComboBox)
        add_form_row(3, "Gender", ctk.CTkComboBox, "Phone Number", ctk.CTkEntry)
        add_form_row(4, "Address", ctk.CTkEntry, "Booking Type", ctk.CTkComboBox)
        add_form_row(5, "Arrival Date", DateEntry, "Departure Date", DateEntry)
        add_form_row(6, "Vehicle No", ctk.CTkEntry, "Attachment", ctk.CTkEntry)

        # Submit button
        submit_button = ctk.CTkButton(
        form_frame,
        text="Update",
        width=140,
        height=36,
        corner_radius=10,
        font=("Segoe UI", 12, "bold"),
        text_color="white",
        fg_color="#2563eb",  # Slightly lighter blue
        hover_color="#1d4ed8",
        command=lambda: self.save_updated_booking(window)
    )

        submit_button.grid(row=7, column=0, columnspan=4, pady=20, sticky="n")

        # Populate fields from booking_data
        if booking_data:
            for label_text, entry in self.update_entries.items():
                key = label_text.lower().replace(" ", "_")
                value = booking_data.get(key, "")

                if isinstance(entry, DateEntry):
                    try:
                        date_obj = datetime.strptime(value, "%Y-%m-%d")
                        entry.set_date(date_obj)
                    except Exception as e:
                        print(f"Failed to parse date for {label_text}: {e}")
                elif isinstance(entry, ctk.CTkComboBox):
                    entry.set(value)
                elif label_text == "Attachment":
                    entry.configure(state="normal")
                    entry.delete(0, "end")
                    entry.insert(0, value if value else "Select file...")  # Insert full path
                    entry.configure(text_color="black" if value else "gray")

                else:
                    entry.configure(state="normal")
                    entry.delete(0, "end")
                    entry.insert(0, value)
                    if label_text == "Booking ID":
                        entry.configure(state="readonly")



    


    def save_updated_booking(self, window):
        try:
            url = "http://127.0.0.1:8000/bookings/update/"

            # Extract raw form values
            booking_id = self.update_entries["Booking ID"].get()
            room_number = self.update_entries["Room Number"].get()
            guest_name = self.update_entries["Guest Name"].get()
            gender = self.update_entries["Gender"].get()
            mode_of_identification = self.update_entries["Mode of Identification"].get()
            identification_number = self.update_entries["Identification Number"].get()
            address = self.update_entries["Address"].get()
            arrival_date = self.update_entries["Arrival Date"].get()
            departure_date = self.update_entries["Departure Date"].get()
            booking_type = self.update_entries["Booking Type"].get()
            phone_number = self.update_entries["Phone Number"].get()
            vehicle_no = self.update_entries["Vehicle No"].get()
            attachment_path = self.update_entries["Attachment"].get()

            # ✅ Convert date to YYYY-MM-DD
            arrival_date = self.update_entries["Arrival Date"].get_date().strftime("%Y-%m-%d")
            departure_date = self.update_entries["Departure Date"].get_date().strftime("%Y-%m-%d")



            if not self.token:
                messagebox.showerror("Authentication Error", "You are not authenticated. Please log in again.")
                return

            headers = {"Authorization": f"Bearer {self.token}"}

            data = {
                "booking_id": booking_id,
                "room_number": room_number,
                "guest_name": guest_name,
                "gender": gender,
                "mode_of_identification": mode_of_identification,
                "identification_number": identification_number,
                "address": address,
                "arrival_date": arrival_date,
                "departure_date": departure_date,
                "booking_type": booking_type,
                "phone_number": phone_number,
                "vehicle_no": vehicle_no,
            }

            files = None

            # Check if user browsed a new file
            if os.path.isfile(attachment_path):  # New file selected via file dialog
                file_name = os.path.basename(attachment_path)
                files = {
                    "attachment": (file_name, open(attachment_path, "rb"), "application/octet-stream")
                }
            else:
                # Use existing attachment path (assume backend handles this logic)
                data["attachment_str"] = attachment_path  # ✅ Matches backend expectation




            response = requests.put(url, data=data, files=files, headers=headers)

            if response.status_code == 200:
                messagebox.showinfo("Success", "Booking updated successfully.")
                window.destroy()  # Close the update popup

                
                #self.refresh_booking_list()
            else:
                messagebox.showerror("Error", f"Failed to update booking:\n{response.text}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while updating the booking:\n{str(e)}")


    

    def view_selected_booking(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("No selection", "Please select a booking to view.")
            return

        booking_data = self.tree.item(selected_item, "values")

        field_names = [
            "Booking ID", "Room No", "Guest Name", "Gender", "Booking Cost", "Arrival",
            "Departure", "Status", "No of Days", "Booking Type", "Phone Number", "Booking Date",
            "Payment Status", "Mode of Identification", "Identification No", "Address", "Vehicle No", "Attachment", "Created by"
        ]

        view_window = ctk.CTkToplevel(self.root)
        view_window.title("Booking Details")
        view_window.geometry("700x500+100+10")
        view_window.configure(fg_color="#ebebeb")
        view_window.grab_set()

        hotel_label = ctk.CTkLabel(
            view_window,
            text=HOTEL_NAME,
            font=("Arial", 17, "bold"),
            text_color="#0f2e4d"
        )
        hotel_label.pack(pady=(10, 0))

        title_label = ctk.CTkLabel(
            view_window,
            text="Guest Details Report",
            font=("Arial", 15, "bold"),
            text_color="#1e3d59"
        )
        title_label.pack(pady=(5, 10))

        content_frame = ctk.CTkFrame(
            view_window,
            fg_color="white",
            border_color="#cccccc",
            border_width=1,
            corner_radius=12
        )
        content_frame.pack(fill="both", expand=False, padx=15, pady=(5, 10))

        content_frame.grid_columnconfigure(0, weight=0)
        content_frame.grid_columnconfigure(1, weight=0)
        content_frame.grid_columnconfigure(2, weight=0)
        content_frame.grid_columnconfigure(3, weight=0)


        rows = []
        attachment_url = None

        for idx, (field, value) in enumerate(zip(field_names, booking_data)):
            col = 0 if idx < 10 else 2
            row = idx % 10

            label_field = ctk.CTkLabel(
                content_frame,
                text=f"{field}:",
                font=("Arial", 12, "bold"),
                text_color="#2c3e50",
                anchor="w"
            )
            label_field.grid(row=row, column=col, sticky="w", padx=(20, 15), pady=3)

            label_value = ctk.CTkLabel(
                content_frame,
                text=str(value),
                font=("Arial", 12),
                text_color="#34495e",
                anchor="w",
                wraplength=240
            )
            label_value.grid(row=row, column=col + 1, sticky="w", padx=(0, 25), pady=3)

            rows.append((field, value))

            if field.lower() == "attachment" and value:
                attachment_url = value

        # --- Display Attachment Image (Top-right corner) ---
        if attachment_url:
            try:
                filename = os.path.basename(attachment_url)
                url = f"http://127.0.0.1:8000/files/attachments/{filename}"
                response = requests.get(url)

                img = Image.open(BytesIO(response.content))
                img = img.convert("RGBA")
                ctk_image = CTkImage(light_image=img, size=(120, 120))

                attachment_preview = ctk.CTkLabel(
                    master=view_window,
                    image=ctk_image,
                    text="",
                    width=120,
                    height=120
                )
                attachment_preview.image = ctk_image  # prevent GC
                attachment_preview.place(x=580, y=20)
            except Exception as e:
                print("")

        # --- Buttons ---
        button_frame = ctk.CTkFrame(view_window, fg_color="white")
        button_frame.pack(pady=(10, 15))

        pdf_button = ctk.CTkButton(
            master=button_frame,
            text="Print Details",
            command=lambda: self.export_booking_to_pdf(rows),
            fg_color="#1e3d59",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=125,
            height=25
        )
        pdf_button.grid(row=0, column=0, padx=5)

        guest_form_btn = ctk.CTkButton(
            master=button_frame,
            text="Guest Form",
            command=lambda: self.export_guest_form(rows),
            fg_color="#1e3d59",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=125,
            height=25
        )
        guest_form_btn.grid(row=0, column=1, padx=5)

        if attachment_url:
            open_attachment_btn = ctk.CTkButton(
                master=button_frame,
                text="Open Attachment",
                command=lambda: self.open_attachment(attachment_url),
                fg_color="#1e3d59",
                text_color="white",
                corner_radius=20,
                font=("Arial", 12, "bold"),
                width=125,
                height=25
            )
            open_attachment_btn.grid(row=0, column=2, padx=5)




    def export_guest_form(self, data):
        # Use NamedTemporaryFile for safer temp file handling
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp:
            filename = temp.name

        doc = SimpleDocTemplate(filename, pagesize=A4,
                                rightMargin=40, leftMargin=40, topMargin=25, bottomMargin=25)

        styles = getSampleStyleSheet()
        elements = []

        # Hotel name (centered and bold)
        hotel_name = Paragraph(f"<para alignment='center'><b>{HOTEL_NAME}</b></para>", styles['Title'])
        elements.append(hotel_name)
        elements.append(Spacer(1, 1))

        # Booking report title (centered)
        title = Paragraph("<para alignment='center'>Guest Details Report</para>", styles['Heading2'])
        elements.append(title)
        elements.append(Spacer(1, 1))

        # Extract booking date if present
        booking_date = ""
        for field, value in data:
            if field.lower() == "booking date":
                booking_date = str(value)
                break

        if booking_date:
            date_paragraph = Paragraph(f"<para alignment='center'><b>Booking Date:</b> {booking_date}</para>", styles['Normal'])
            elements.append(date_paragraph)
            elements.append(Spacer(1, 2))

        # Format data into table
        table_data = [[Paragraph(f"<b>{field}</b>", styles['Normal']), Paragraph(str(value), styles['Normal'])] for field, value in data]

        table = Table(table_data, colWidths=[120, 330])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 10))

        # Legal preamble before rules and directives
        preamble_text = """
        <b>Guest Acknowledgement and Agreement</b><br/><br/>
        I, _____________________________________, hereby acknowledge that I have carefully read, understood, and agreed to abide by the rules and directives outlined below. 
        I understand that failure to comply with any of the stated rules may result in penalties, fines, or eviction from the hotel premises without refund. 
        I accept full responsibility for any violations and understand that I will be held liable as specified herein.<br/><br/>
        """

        preamble_paragraph = Paragraph(preamble_text, styles['Normal'])
        elements.append(preamble_paragraph)

        # Rules and directives section
        rules_text = """
        <b>RULES AND DIRECTIVES</b><br/>
        X No smoking of any kind or intake of hard drugs in the room or toilet. Fine is N200,000.<br/>
        X Guests are not allowed to bring in food and drinks from outside into the hotel premises.<br/>
        X Smoking of cigarette is only allowed by the Pool Bar area.<br/>
        X All guests are to drop their keycards whenever they are leaving the hotel premises.<br/>
        X Misplacement of room keycard attracts a fine of N5,000 for immediate replacement.<br/>
        X Destroying or staining towels, bedsheets, rugs, and especially wallpaper in the room or any area in the hotel will attract immediate replacement.<br/><br/>
        <b>NOTE:</b> Guests that go contrary to the above stated rules and directives will be evicted without refund.
        """

        rules_paragraph = Paragraph(rules_text, styles['Normal'])
        elements.append(rules_paragraph)
        elements.append(Spacer(1, 25))

        # Add signature lines
        signature_table = Table([
            ["Guest Signature ________________________", "Receptionist Signature ____________________"]
        ], colWidths=[260, 260])
        elements.append(signature_table)

        # Build PDF
        doc.build(elements)

        # Open PDF in default viewer
        webbrowser.open_new(f'file://{os.path.abspath(filename)}')
        


    def open_attachment(self, attachment_path):
        import os
        import webbrowser

        if not attachment_path:
            messagebox.showerror("Attachment Error", "No attachment found.")
            return

        filename = os.path.basename(attachment_path)
        url = f"http://127.0.0.1:8000/files/attachments/{filename}"
        try:
            webbrowser.open(url)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open attachment: {e}")




    
    


    def export_booking_to_pdf(self, data):
        # Use NamedTemporaryFile for safer temp file handling
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp:
            filename = temp.name

        doc = SimpleDocTemplate(filename, pagesize=A4,
                                rightMargin=40, leftMargin=40, topMargin=25, bottomMargin=25)

        styles = getSampleStyleSheet()
        elements = []

        # Hotel name (centered and bold)
        hotel_name = Paragraph(f"<para alignment='center'><b>{HOTEL_NAME}</b></para>", styles['Title'])
        elements.append(hotel_name)
        elements.append(Spacer(1, 1))

        # Booking report title (centered)
        title = Paragraph("<para alignment='center'>Guest Details Report</para>", styles['Heading2'])
        elements.append(title)
        elements.append(Spacer(1, 1))

        # Extract booking date and attachment
        booking_date = ""
        attachment_data = None

        filtered_data = []
        for field, value in data:
            if field.lower() == "booking date":
                booking_date = str(value)
            elif field.lower() == "attachment":
                if isinstance(value, str):
                    # 🔁 Convert public URL path to local path if necessary
                    if value.startswith("/uploads/attachments/"):
                        local_path = os.path.join("uploads", "attachments", os.path.basename(value))
                    else:
                        local_path = value

                    if os.path.exists(local_path):
                        with open(local_path, "rb") as f:
                            attachment_data = f.read()
                elif isinstance(value, bytes):
                    attachment_data = value

            else:
                filtered_data.append((field, value))

        if booking_date:
            date_paragraph = Paragraph(f"<para alignment='center'><b>Booking Date:</b> {booking_date}</para>", styles['Normal'])
            elements.append(date_paragraph)
            elements.append(Spacer(1, 2))

        # Format data into table (excluding image)
        table_data = [[Paragraph(f"<b>{field}</b>", styles['Normal']), Paragraph(str(value), styles['Normal'])] for field, value in filtered_data]

        table = Table(table_data, colWidths=[120, 330])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 15))

        # Add attachment image if available
        # Display attachment image
        if attachment_data:
            try:
                image_stream = BytesIO(attachment_data)
                img = Image(image_stream, width=300, height=250)  # ✅ Corrected line
                elements.append(Spacer(1, 12))
                elements.append(Paragraph("Attachment Photo:", styles['Heading2']))
                elements.append(img)
            except Exception as e:
                print("Image error:", e)
                elements.append(Paragraph("Could not display the attachment image.", styles["Normal"]))


        # Add signature lines
        #signature_table = Table([
            #["Guest Signature ________________________", "Receptionist Signature ____________________"]
        #], colWidths=[260, 260])
        #elements.append(signature_table)

        # Build PDF
        doc.build(elements)

        # Open PDF in default viewer
        webbrowser.open_new(f'file://{os.path.abspath(filename)}')


        

    def fetch_bookings(self, start_date_entry, end_date_entry):
        """Fetch bookings from the API and populate the table, while calculating total booking cost."""
        api_url ="http://127.0.0.1:8000/bookings/list"  # Ensure correct endpoint
        params = {
            "start_date": start_date_entry.get_date().strftime("%Y-%m-%d"),
            "end_date": end_date_entry.get_date().strftime("%Y-%m-%d"),
        }
        headers = {"Authorization": f"Bearer {self.token}"}

        try:
            response = requests.get(api_url, params=params, headers=headers)
            if response.status_code == 200:
                data = response.json()
                # print("API Response:", data)  # Debugging output

                if isinstance(data, dict) and "bookings" in data:
                    bookings = data["bookings"]
                    total_booking_cost = data.get("total_booking_cost", 0)  # Get total cost from API
                    total_entries = data.get("total_entries", len(bookings))  # Fallback in case key missing
                else:
                    messagebox.showerror("Error", "Unexpected API response format")
                    return

                # Check if bookings list is empty
                if not bookings:
                    self.total_booking_cost_label.config(text="Total Booking Cost: 0.00")  # Reset label
                    self.total_entries_label.config(text="Total Entries: 0")  # Clear label too
                    messagebox.showinfo("No Results", "No bookings found for the selected filters.")
                    return

                self.tree.delete(*self.tree.get_children())  # Clear table

                for booking in bookings:
                    self.tree.insert("", "end", values=(
                        booking.get("id", ""),
                        booking.get("room_number", ""),
                        booking.get("guest_name", ""),
                        booking.get("gender", ""),
                        f"{float(booking.get('booking_cost', 0)) :,.2f}",
                        booking.get("arrival_date", ""),
                        booking.get("departure_date", ""),
                        booking.get("status", ""),
                        booking.get("number_of_days", ""),
                        booking.get("booking_type", ""),
                        booking.get("phone_number", ""),
                        booking.get("booking_date", ""),
                        booking.get("payment_status", ""), 
                        booking.get("mode_of_identification", ""), 
                        booking.get("identification_number", ""),  
                        booking.get("address", ""),                 
                        booking.get("vehicle_no", ""),
                        booking.get("attachment", ""),
                        booking.get("created_by", ""),

                    ))

                # Apply grid effect after inserting data
                self.apply_grid_effect()

                # Display total booking cost
                self.total_booking_cost_label.config(
                    text=f"Total Booking Cost: {total_booking_cost:,.2f}"
                )

                self.total_entries_label.config(text=f"Total Entries: {len(bookings)}")

            
            
            else:
                messagebox.showerror("Error", response.json().get("detail", "Failed to retrieve bookings."))

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error", f"Request failed: {e}")


    def clear_right_frame(self):
        for widget in self.right_frame.winfo_children():
            widget.pack_forget()

    
    
    
    
    def list_bookings_by_status(self):
        """Displays the List Bookings by Status UI."""
        self.current_view = "status_search"
        
        self.clear_right_frame()  # Ensure old UI elements are removed

        # Create a new frame for the table with scrollable functionality
        frame = tk.Frame(self.right_frame, bg="#D6D8DA", padx=10, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="List Bookings by Status", font=("Arial", 14, "bold"), bg="#D6D8DA").pack(pady=10)

        # Filter Frame
        filter_frame = tk.Frame(frame, bg="#ffffff")
        filter_frame.pack(pady=5)

        # Status Dropdown
       # Status Dropdown
        tk.Label(filter_frame, text="Status:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=0, padx=5, pady=5)

        status_options = ["checked-in", "reserved", "checked-out", "cancelled", "complimentary"]
        self.status_var = tk.StringVar(value=status_options[0])  # Default selection

        status_menu = ttk.Combobox(filter_frame, textvariable=self.status_var, values=status_options, state="readonly")
        status_menu.grid(row=0, column=1, padx=5, pady=5)

        # Bind the selection event to a function that updates self.status_var
        def on_status_change(event):
            #print("Selected Status:", self.status_var.get())  # Debugging: Check what is selected
            self.status_var.set(status_menu.get())  # Ensure value updates

        status_menu.bind("<<ComboboxSelected>>", on_status_change)  # Event binding


        # Start Date
        tk.Label(filter_frame, text="Start Date:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=2, padx=5, pady=5)
        self.start_date = DateEntry(filter_frame, font=("Arial", 11))
        self.start_date.grid(row=0, column=3, padx=5, pady=5)

        # End Date
        tk.Label(filter_frame, text="End Date:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=4, padx=5, pady=5)
        self.end_date = DateEntry(filter_frame, font=("Arial", 11))
        self.end_date.grid(row=0, column=5, padx=5, pady=5)

        # Fetch Button
        fetch_btn = ttk.Button(filter_frame, text="Fetch Bookings", command=self.fetch_bookings_by_status)
        fetch_btn.grid(row=0, column=6, padx=10, pady=5)

        # Table Frame
        table_frame = tk.Frame(frame, bg="#ffffff")
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("ID", "Room No", "Guest Name", "Gender", "Booking Cost", "Arrival",
            "Departure", "Status", "No of Days", "Booking Type", "Phone Number", "Booking Date",
            "Payment Status", "Mode of Identification", "Identification No", "Address", "Vehicle No", "Attachment", "Created by")


        # ✅ Prevent recreation of table on every call
        if hasattr(self, "tree"):
            self.tree.destroy()

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=70, anchor="center")
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscroll=y_scroll.set)
        
        x_scroll = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        x_scroll.pack(fill=tk.X)
        self.tree.configure(xscroll=x_scroll.set)

        # --- View Booking Button using CustomTkinter ---
        view_button = ctk.CTkButton(
            frame,
            text="View Booking",
            corner_radius=20,
            command=self.view_selected_guest_Status
        )
        view_button.pack(pady=3)


        # ✅ Add Label for Total Booking Cost at the Bottom
        self.total_cost_label = tk.Label(frame, text="Total Booking Cost: 0.00", font=("Arial", 12, "bold"), bg="#D6D8DA", fg="blue")
        self.total_cost_label.pack(pady=10)  # Placed at the bottom

        self.total_entries_label = tk.Label(frame, text="Total Entries: 0", font=("Arial", 12, "bold"), bg="#D6D8DA", fg="green")
        self.total_entries_label.pack()

        
    def view_selected_guest_Status(self):
        #selected_item = self.tree.focus()

        if hasattr(self, "tree"):
            selected_item = self.tree.focus()
        else:
            messagebox.showerror("Error", "Booking list not available.")
            return


        booking_data = self.tree.item(selected_item)["values"]


        field_names = [
            "ID", "Room", "Guest Name", "Gender", "Booking Cost", "Arrival",
            "Departure", "Status", "Number of Days", "Booking Type", "Phone Number", "Booking Date",
            "Payment Status", "Mode of Identification", "Identification Number",
            "Address", "Vehicle No", "Attachment", "Created By"
        ]

        popup = ctk.CTkToplevel(self.root)
        popup.title("Guest Booking Details")
        popup.geometry("700x500+100+20")
        popup.configure(fg_color="#ebebeb")
        popup.grab_set()

        hotel_label = ctk.CTkLabel(
            popup, text=HOTEL_NAME, font=("Arial", 17, "bold"), text_color="#0f2e4d"
        )
        hotel_label.pack(pady=(10, 0))

        title_label = ctk.CTkLabel(
            popup,
            text="Guest Booking Details",
            font=("Arial", 15, "bold"),
            text_color="#1e3d59"
        )
        title_label.pack(pady=(5, 10))

        content_frame = ctk.CTkFrame(
            popup,
            fg_color="white",
            border_color="#cccccc",
            border_width=1,
            corner_radius=12
        )
        content_frame.pack(fill="both", expand=False, padx=15, pady=(5, 10))

        content_frame.grid_columnconfigure(0, weight=0)
        content_frame.grid_columnconfigure(1, weight=0)
        content_frame.grid_columnconfigure(2, weight=0)
        content_frame.grid_columnconfigure(3, weight=0)

        rows = []
        attachment_url = None

        for idx, (field, value) in enumerate(zip(field_names, booking_data)):
            col = 0 if idx < 10 else 2
            row = idx % 10

            label_field = ctk.CTkLabel(
                content_frame,
                text=f"{field}:",
                font=("Arial", 12, "bold"),
                text_color="#2c3e50",
                anchor="w"
            )
            label_field.grid(row=row, column=col, sticky="w", padx=(20, 15), pady=3)

            label_value = ctk.CTkLabel(
                content_frame,
                text=str(value),
                font=("Arial", 12),
                text_color="#34495e",
                anchor="w",
                wraplength=240
            )
            label_value.grid(row=row, column=col + 1, sticky="w", padx=(0, 25), pady=3)

            rows.append((field, value))

            if field.lower() == "attachment" and value:
                attachment_url = value

        # --- Optional: Display Attachment as Image Preview (top-right corner) ---
        if attachment_url:
            try:
                filename = os.path.basename(attachment_url)
                url = f"http://127.0.0.1:8000/files/attachments/{filename}"
                response = requests.get(url)

                img = Image.open(BytesIO(response.content))
                img = img.convert("RGBA")
                ctk_image = CTkImage(light_image=img, size=(120, 120))

                attachment_preview = ctk.CTkLabel(
                    master=popup,
                    image=ctk_image,
                    text="",
                    width=120,
                    height=120
                )
                attachment_preview.image = ctk_image
                attachment_preview.place(x=600, y=20)
            except Exception as e:
                print("")

        # --- Action Buttons ---
        button_frame = ctk.CTkFrame(popup, fg_color="white")
        button_frame.pack(pady=(10, 15))

        pdf_button = ctk.CTkButton(
            master=button_frame,
            text="Print Details",
            command=lambda: self.export_booking_to_pdf(rows),
            fg_color="#1e3d59",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=120,
            height=25
        )
        pdf_button.grid(row=0, column=0, padx=5)

        guest_form_btn = ctk.CTkButton(
            master=button_frame,
            text="Guest Form",
            command=lambda: self.export_guest_form(rows),
            fg_color="#1e3d59",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=120,
            height=25
        )
        guest_form_btn.grid(row=0, column=1, padx=5)

        close_btn = ctk.CTkButton(
            master=button_frame,
            text="Close",
            command=popup.destroy,
            fg_color="#cc0000",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=120,
            height=25
        )
        close_btn.grid(row=0, column=2, padx=5)
    



    def fetch_bookings_by_status(self):
        """Fetch bookings based on status and date filters."""
        api_url = "http://127.0.0.1:8000/bookings/status"

        selected_status = self.status_var.get().strip().lower()  # Ensure correct status retrieval

        # ✅ Debugging: Print the selected status before sending
        #print(f"Selected Status from Dropdown: '{selected_status}'")

        params = {
            "status": selected_status,  # Ensure correct status is passed
            "start_date": self.start_date.get_date().strftime("%Y-%m-%d"),
            "end_date": self.end_date.get_date().strftime("%Y-%m-%d"),
        }

        headers = {"Authorization": f"Bearer {self.token}"}

        try:
            response = requests.get(api_url, params=params, headers=headers)
            data = response.json()

            # ✅ Debugging: Print the API response
            #print("API Response:", data)

            if response.status_code == 200:
                if "bookings" in data and isinstance(data["bookings"], list):
                    bookings = data["bookings"]

                    self.tree.delete(*self.tree.get_children())  # Clear previous data

                    total_cost = 0  # Initialize total booking cost

                    if bookings:
                        for booking in bookings:
                            is_canceled = booking.get("status", "").lower() == "cancelled"
                            tag = "cancelled" if is_canceled else "normal"

                            booking_cost = float(booking.get("booking_cost", 0))
                            total_cost += booking_cost

                            self.tree.insert("", "end", values=(
                                booking.get("id", ""),
                                booking.get("room_number", ""),
                                booking.get("guest_name", ""),
                                booking.get("gender", ""),
                                f"{booking_cost:,.2f}",
                                booking.get("arrival_date", ""),
                                booking.get("departure_date", ""),
                                booking.get("status", ""),
                                booking.get("number_of_days", ""),
                                booking.get("booking_type", ""),
                                booking.get("phone_number", ""),
                                booking.get("booking_date", ""),
                                booking.get("payment_status", ""), 
                                booking.get("mode_of_identification", ""),
                                booking.get("identification_number", ""),  
                                booking.get("address", ""),                               
                                booking.get("vehicle_no", ""),
                                booking.get("attachment", ""),
                                booking.get("created_by", ""),

                            ), tags=(tag,))

                        # Apply grid effect after inserting data
                        self.apply_grid_effect()


                        self.tree.tag_configure("cancelled", foreground="red")
                        self.tree.tag_configure("normal", foreground="black")
                        self.total_cost_label.config(text=f"Total Booking Cost: {total_cost:,.2f}")
                        self.total_entries_label.config(text=f"Total Entries: {len(bookings)}")


                        
                    else:
                        self.tree.delete(*self.tree.get_children())
                        self.total_cost_label.config(text="Total Booking Cost: ₦0.00")
                        messagebox.showinfo("No Results", "No bookings found for the selected filters.")

                elif "message" in data:
                    messagebox.showinfo("Info", data["message"])
                    self.tree.delete(*self.tree.get_children())
                    self.total_cost_label.config(text="Total Booking Cost: ₦0.00")

            else:
                messagebox.showerror("Error", response.json().get("detail", "Failed to retrieve bookings."))

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error", f"Request failed: {e}")



            
    #Search Booking by Guest Name
    def search_booking_by_guest_name(self):
        self.current_view = "guest_search"

        self.clear_right_frame()
        
        frame = tk.Frame(self.right_frame, bg="#D6D8DA", padx=10, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(frame, text="Search Booking by Guest Name", font=("Arial", 14, "bold"), bg="#D6D8DA").pack(pady=10)
        
        search_frame = tk.Frame(frame, bg="#ffffff")
        search_frame.pack(pady=5)
        
        tk.Label(search_frame, text="Guest Name:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=0, padx=5, pady=5)
        self.search_entry = tk.Entry(search_frame, font=("Arial", 11))
        self.search_entry.grid(row=0, column=1, padx=5, pady=5)
        
        search_btn = ttk.Button(
            search_frame, text="Search", command=self.fetch_booking_by_guest_name
        )
        search_btn.grid(row=0, column=2, padx=10, pady=5)
        
        table_frame = tk.Frame(frame, bg="#ffffff")
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ("ID", "Room No", "Guest Name", "Gender", "Booking Cost", "Arrival",
            "Departure", "Status", "No of Days", "Booking Type", "Phone Number", "Booking Date",
            "Payment Status", "Mode of Identification", "Identification No", "Address", "Vehicle No", "Attachment", "Created by")

        self.search_tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns:
            self.search_tree.heading(col, text=col)
            self.search_tree.column(col, width=70, anchor="center")
        
        self.search_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.search_tree.yview)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.search_tree.configure(yscroll=y_scroll.set)
        
        x_scroll = ttk.Scrollbar(frame, orient="horizontal", command=self.search_tree.xview)
        x_scroll.pack(fill=tk.X)
        self.search_tree.configure(xscroll=x_scroll.set)

      # --- View Booking Button using CustomTkinter ---
        view_button = ctk.CTkButton(
            frame,
            text="View Booking",
            corner_radius=20,
            command=self.view_selected_guest_booking
        )
        view_button.pack(pady=3)




        # Create a horizontal frame for the summary labels
        summary_frame = tk.Frame(frame, bg="#D6D8DA")
        summary_frame.pack(fill=tk.X, pady=10)

        self.total_entries_label = tk.Label(summary_frame, text="", font=("Arial", 12, "bold"), fg="blue", bg="#D6D8DA")
        self.total_entries_label.pack(side=tk.LEFT, padx=10)

        self.total_cost_label = tk.Label(summary_frame, text="", font=("Arial", 12, "bold"), fg="red", bg="#D6D8DA")
        self.total_cost_label.pack(side=tk.LEFT, padx=20)


    def view_selected_guest_booking(self):
        selected_item = self.search_tree.focus()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a booking to view.")
            return

        booking_data = self.search_tree.item(selected_item)["values"]

        field_names = [
            "ID", "Room", "Guest Name", "Gender", "Booking Cost", "Arrival",
            "Departure", "Status", "Number of Days", "Booking Type", "Phone Number", "Booking Date",
            "Payment Status", "Mode of Identification", "Identification Number",
            "Address", "Vehicle No", "Attachment", "Created By"
        ]

        popup = ctk.CTkToplevel(self.root)
        popup.title("Guest Booking Details")
        popup.geometry("700x500+100+20")
        popup.configure(fg_color="#ebebeb")
        popup.grab_set()

        hotel_label = ctk.CTkLabel(
            popup, text=HOTEL_NAME, font=("Arial", 17, "bold"), text_color="#0f2e4d"
        )
        hotel_label.pack(pady=(10, 0))

        title_label = ctk.CTkLabel(
            popup,
            text="Guest Booking Details",
            font=("Arial", 15, "bold"),
            text_color="#1e3d59"
        )
        title_label.pack(pady=(5, 10))

        content_frame = ctk.CTkFrame(
            popup,
            fg_color="white",
            border_color="#cccccc",
            border_width=1,
            corner_radius=12
        )
        content_frame.pack(fill="both", expand=False, padx=15, pady=(5, 10))

        content_frame.grid_columnconfigure(0, weight=0)
        content_frame.grid_columnconfigure(1, weight=0)
        content_frame.grid_columnconfigure(2, weight=0)
        content_frame.grid_columnconfigure(3, weight=0)

        rows = []
        attachment_url = None

        for idx, (field, value) in enumerate(zip(field_names, booking_data)):
            col = 0 if idx < 10 else 2
            row = idx % 10

            label_field = ctk.CTkLabel(
                content_frame,
                text=f"{field}:",
                font=("Arial", 12, "bold"),
                text_color="#2c3e50",
                anchor="w"
            )
            label_field.grid(row=row, column=col, sticky="w", padx=(20, 15), pady=3)

            label_value = ctk.CTkLabel(
                content_frame,
                text=str(value),
                font=("Arial", 12),
                text_color="#34495e",
                anchor="w",
                wraplength=240
            )
            label_value.grid(row=row, column=col + 1, sticky="w", padx=(0, 25), pady=3)

            rows.append((field, value))

            if field.lower() == "attachment" and value:
                attachment_url = value

        # --- Optional: Display Attachment as Image Preview (top-right corner) ---
        if attachment_url:
            try:
                filename = os.path.basename(attachment_url)
                url = f"http://127.0.0.1:8000/files/attachments/{filename}"
                response = requests.get(url)

                img = Image.open(BytesIO(response.content))
                img = img.convert("RGBA")
                ctk_image = CTkImage(light_image=img, size=(120, 120))

                attachment_preview = ctk.CTkLabel(
                    master=popup,
                    image=ctk_image,
                    text="",
                    width=120,
                    height=120
                )
                attachment_preview.image = ctk_image
                attachment_preview.place(x=600, y=20)
            except Exception as e:
                print("")

        # --- Action Buttons ---
        button_frame = ctk.CTkFrame(popup, fg_color="white")
        button_frame.pack(pady=(10, 15))

        pdf_button = ctk.CTkButton(
            master=button_frame,
            text="Print Details",
            command=lambda: self.export_booking_to_pdf(rows),
            fg_color="#1e3d59",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=120,
            height=25
        )
        pdf_button.grid(row=0, column=0, padx=5)

        guest_form_btn = ctk.CTkButton(
            master=button_frame,
            text="Guest Form",
            command=lambda: self.export_guest_form(rows),
            fg_color="#1e3d59",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=120,
            height=25
        )
        guest_form_btn.grid(row=0, column=1, padx=5)

        close_btn = ctk.CTkButton(
            master=button_frame,
            text="Close",
            command=popup.destroy,
            fg_color="#cc0000",
            text_color="white",
            corner_radius=20,
            font=("Arial", 12, "bold"),
            width=120,
            height=25
        )
        close_btn.grid(row=0, column=2, padx=5)





    
    def fetch_booking_by_guest_name(self):
        guest_name = self.search_entry.get().strip()
        if not guest_name:
            messagebox.showerror("Error", "Please enter a guest name to search.")
            return

        
         
        api_url ="http://127.0.0.1:8000/bookings/search"
        params = {"guest_name": guest_name}
        headers = {"Authorization": f"Bearer {self.token}"}
        
        try:
            response = requests.get(api_url, params=params, headers=headers)
            if response.status_code == 200:
                data = response.json()
                bookings = data.get("bookings", [])
                
                self.search_tree.delete(*self.search_tree.get_children())
                
                for booking in bookings:
                    self.search_tree.insert("", "end", values=(
                        booking.get("id", ""),
                        booking.get("room_number", ""),
                        booking.get("guest_name", ""),
                        booking.get("gender", ""),
                         f"{float(booking.get('booking_cost', 0)) :,.2f}",  # Format booking_cost
                        booking.get("arrival_date", ""),
                        booking.get("departure_date", ""),
                        booking.get("status", ""),
                        booking.get("number_of_days", ""),
                        booking.get("booking_type", ""),
                        booking.get("phone_number", ""),
                        booking.get("booking_date", ""),
                        booking.get("payment_status", ""),
                        booking.get("mode_of_identification", ""),
                        booking.get("identification_number", ""),  
                        booking.get("address", ""),                               
                        booking.get("vehicle_no", ""),
                        booking.get("attachment", ""),
                        booking.get("created_by", ""),


                       
                    ))
            
                # Apply grid effect after inserting data
                self.apply_grid_effect(self.search_tree)

                # ✅ Show total entries and booking cost excluding cancelled/complimentary
                total_entries = data.get("total_bookings", 0)
                total_cost = data.get("total_booking_cost", 0.0)

                self.total_entries_label.config(text=f"Total Entries: {total_entries:,}")
                self.total_cost_label.config(text=f"Total Booking Cost: ₦{total_cost:,.2f}")

                

    
            else:
                messagebox.showinfo("No result", response.json().get("detail", "No bookings found."))
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error", f"Request failed: {e}")

    
     
    def search_booking_by_room(self):
        self.current_view = "room_search"

        self.clear_right_frame()

        frame = tk.Frame(self.right_frame, bg="#D6D8DA", padx=10, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="Search Booking by Room Number", font=("Arial", 14, "bold"), bg="#D6D8DA").pack(pady=10)

        # Centered Frame for Search Inputs
        search_frame = tk.Frame(frame, bg="#ffffff")
        search_frame.pack(pady=10)

        # Grid Configuration for Centralization
        search_frame.grid_columnconfigure(0, weight=1)
        search_frame.grid_columnconfigure(1, weight=1)
        search_frame.grid_columnconfigure(2, weight=1)
        search_frame.grid_columnconfigure(3, weight=1)
        search_frame.grid_columnconfigure(4, weight=1)
        search_frame.grid_columnconfigure(5, weight=1)
        search_frame.grid_columnconfigure(6, weight=1)

        # Room Number Entry
        tk.Label(search_frame, text="Room Number:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.room_number_entry = tk.Entry(search_frame, font=("Arial", 11), width=12)
        self.room_number_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # Start Date
        tk.Label(search_frame, text="Start Date:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=2, padx=5, pady=5, sticky="e")
        self.start_date_entry = DateEntry(search_frame, font=("Arial", 11), width=12, background="darkblue", foreground="white", borderwidth=2)
        self.start_date_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        # End Date
        tk.Label(search_frame, text="End Date:", font=("Arial", 11), bg="#ffffff").grid(row=0, column=4, padx=5, pady=5, sticky="e")
        self.end_date_entry = DateEntry(search_frame, font=("Arial", 11), width=12, background="darkblue", foreground="white", borderwidth=2)
        self.end_date_entry.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        # Search Button
        search_btn = ttk.Button(search_frame, text="Search", command=self.fetch_booking_by_room)
        search_btn.grid(row=0, column=6, padx=10, pady=5, sticky="w")

        # Table Frame for results
        table_frame = tk.Frame(frame, bg="#ffffff")
        table_frame.pack(fill=tk.BOTH, expand=True)

        # Define table columns
        columns = ("ID", "Room", "Guest Name", "Gender", "Booking Cost", "Arrival",
            "Departure", "Status", "Number of Days", "Booking Type", "Phone Number", "Booking Date",
            "Payment Status", "Mode of Identification", "Identification Number",
            "Address", "Vehicle No", "Attachment", "Created By")

        self.search_tree = ttk.Treeview(table_frame, columns=columns, show="headings")

        # Configure table column headings
        for col in columns:
            self.search_tree.heading(col, text=col)
            self.search_tree.column(col, width=70, anchor="center")

        self.search_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Scrollbars
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.search_tree.yview)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.search_tree.configure(yscroll=y_scroll.set)

        x_scroll = ttk.Scrollbar(frame, orient="horizontal", command=self.search_tree.xview)
        x_scroll.pack(fill=tk.X)
        self.search_tree.configure(xscroll=x_scroll.set)


        # --- View Booking Button using CustomTkinter ---
        view_button = ctk.CTkButton(
            frame,
            text="View Booking",
            corner_radius=20,
            command=self.view_selected_guest_booking
        )
        view_button.pack(pady=3)

        

# Frame to hold totals and view button
        totals_frame = tk.Frame(frame, bg="#D6D8DA")
        totals_frame.pack(pady=10, anchor="w")  # Left-align

        # Total Entries Label (Blue)
        self.total_entries_label = tk.Label(totals_frame, text="Total Entries: 0", fg="blue", font=("Arial", 12, "bold"), bg="#D6D8DA")
        self.total_entries_label.pack(side="left", padx=(0, 10))

        # Total Booking Cost Label (Red)
        self.total_label = tk.Label(totals_frame, text="Total Booking Cost: 0.00", fg="red", font=("Arial", 12, "bold"), bg="#D6D8DA")
        self.total_label.pack(side="left", padx=(0, 10))



    



    def fetch_booking_by_room(self):
        room_number = self.room_number_entry.get().strip()

        if not room_number:
            messagebox.showerror("Error", "Please enter a room number.")
            return

        try:
            start_date = self.start_date_entry.get_date()
            end_date = self.end_date_entry.get_date()

            if not start_date or not end_date:
                messagebox.showerror("Error", "Please select both start and end dates.")
                return

            formatted_start_date = start_date.strftime("%Y-%m-%d")
            formatted_end_date = end_date.strftime("%Y-%m-%d")

            api_url = f"http://127.0.0.1:8000/bookings/room/{room_number}"
            params = {"start_date": formatted_start_date, "end_date": formatted_end_date}
            headers = {"Authorization": f"Bearer {self.token}"}

            response = requests.get(api_url, params=params, headers=headers)
            response_data = response.json()

            total_cost = 0.0
            total_entries = 0

            if response.status_code == 200:
                if "bookings" in response_data and response_data["bookings"]:
                    self.search_tree.delete(*self.search_tree.get_children())  # Clear table

                    for booking in response_data["bookings"]:
                        status = booking.get("status", "").lower()
                        booking_type = booking.get("booking_type", "").lower()

                        # Count all bookings regardless of status or type
                        total_entries += 1

                        # For total cost, exclude cancelled and complimentary bookings
                        if status != "cancelled" and booking_type != "complimentary":
                            try:
                                cost = float(booking.get("booking_cost", 0))
                                total_cost += cost
                            except ValueError:
                                pass

                        # Add all bookings to the table
                        self.search_tree.insert("", "end", values=(
                            booking.get("id", ""),
                            booking.get("room_number", ""),
                            booking.get("guest_name", ""),
                            booking.get("gender", ""),
                            f"{float(booking.get('booking_cost', 0)) :,.2f}",
                            booking.get("arrival_date", ""),
                            booking.get("departure_date", ""),
                            booking.get("status", ""),
                            booking.get("number_of_days", ""),
                            booking.get("booking_type", ""),
                            booking.get("phone_number", ""),
                            booking.get("booking_date", ""),
                            booking.get("payment_status", ""),
                            booking.get("mode_of_identification", ""),
                            booking.get("identification_number", ""),
                            booking.get("address", ""),
                            booking.get("vehicle_no", ""),
                            booking.get("attachment", ""),
                            booking.get("created_by", ""),
                        ))

                    # Update totals with color styling
                    self.total_entries_label.config(
                        text=f"Total Entries: {total_entries}",
                        fg="blue"
                    )
                    self.total_label.config(
                        text=f"Total Booking Cost: {total_cost:,.2f}",
                        fg="red"
                    )

                    # Apply grid effect after inserting data
                    self.apply_grid_effect(self.search_tree)

                    
    
                else:
                    messagebox.showinfo("No Results", f"No bookings found for Room {room_number} between {formatted_start_date} and {formatted_end_date}.")
            else:
                error_message = response_data.get("detail", "Failed to retrieve bookings.")
                messagebox.showerror("Error", error_message)

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error", f"Request failed: {e}")

        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error: {e}")



    
    def update_subheading(self, text, command):
        self.subheading_label.config(text=text)
        command()

    


    
#from CTkMessagebox import CTkMessagebox


    def guest_checkout(self):
        """Open a compact, professional pop-up window for guest checkout."""
        self.clear_right_frame()

        # 🔷 Create the popup window
        checkout_window = ctk.CTkToplevel(self.root)
        checkout_window.title("Guest Checkout")
        checkout_window.geometry("360x210")
        checkout_window.resizable(False, False)
        checkout_window.configure(fg_color="#f2f2f2")  # Light gray background

        # Center the window
        screen_width = checkout_window.winfo_screenwidth()
        screen_height = checkout_window.winfo_screenheight()
        x_coordinate = (screen_width - 360) // 2
        y_coordinate = (screen_height - 200) // 2
        checkout_window.geometry(f"360x210+{x_coordinate}+{y_coordinate}")

        checkout_window.transient(self.root)
        checkout_window.grab_set()

        # 🔹 Header with dark background
        header_frame = ctk.CTkFrame(checkout_window, fg_color="#1f2937", corner_radius=10)
        header_frame.pack(fill="x", padx=20, pady=(15, 10))

        header_label = ctk.CTkLabel(
            header_frame,
            text="Guest Checkout",
            font=("Segoe UI", 15, "bold"),
            text_color="white"
        )
        header_label.pack(pady=8)

        # 🔸 Content frame
        content_frame = ctk.CTkFrame(checkout_window, fg_color="white", corner_radius=12)
        content_frame.pack(fill="both", expand=True, padx=20, pady=5)

        # Room number input
        room_label = ctk.CTkLabel(
            content_frame,
            text="Room Number",
            font=("Segoe UI", 12),
            text_color="#1f2937"
        )
        room_label.pack(anchor="w", padx=15, pady=(10, 2))

        self.room_number_entry = ctk.CTkEntry(
            content_frame,
            placeholder_text="enter room number",
            font=("Segoe UI", 12),
            height=32,
            corner_radius=10,
            width=220
        )
        self.room_number_entry.pack(pady=(0, 10))

        # 🔘 Compact checkout button
        submit_btn = ctk.CTkButton(
            content_frame,
            text="Checkout",
            command=lambda: self.submit_guest_checkout(checkout_window),
            font=("Segoe UI", 12, "bold"),
            fg_color="#3B82F6",
            hover_color="#2563EB",
            text_color="white",
            corner_radius=20,
            height=34,
            width=140
        )
        submit_btn.pack(pady=(5, 15))




    def submit_guest_checkout(self, checkout_window):
        """Sends a request to checkout the guest by room number."""
        try:
            room_number = self.room_number_entry.get().strip()

            if not room_number:
                CTkMessagebox(title="Error", message="Please enter a valid room number.", icon="cancel")
                return

            api_url = f"http://127.0.0.1:8000/bookings/{room_number}/"
            headers = {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"}

            response = requests.put(api_url, headers=headers)

            if response.status_code == 200:
                checkout_window.grab_release()  # Release grab before showing message
                # Show success message and wait for user to acknowledge
                msgbox = CTkMessagebox(title="Success", message=f"Guest successfully checked out from room {room_number}!",
                                    icon="check", option_1="OK")
                if msgbox.get() == "OK":  # Check if user clicked "OK"
                    checkout_window.destroy()  # Close the window

            else:
                CTkMessagebox(title="Error", message=response.json().get("detail", "Checkout failed."), icon="warning")

        except requests.exceptions.RequestException as e:
            CTkMessagebox(title="Error", message=f"Request failed: {e}", icon="cancel")

        
            
            
            
    def update_subheading(self, text, command):
        self.subheading_label.config(text=text)
        command()

    



    def cancel_booking(self):
        self.clear_right_frame()

        # Open pop-up window
        cancel_window = tk.Toplevel(self.root)
        cancel_window.title("Cancel Booking")
        cancel_window.configure(bg="#2c3e50")

        # Window size and position
        window_width = 470
        window_height = 260
        screen_width = cancel_window.winfo_screenwidth()
        screen_height = cancel_window.winfo_screenheight()
        x_coordinate = (screen_width - window_width) // 2
        y_coordinate = (screen_height - window_height) // 2
        cancel_window.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

        cancel_window.transient(self.root)
        cancel_window.grab_set()

        # === Header Frame ===
        header_frame = tk.Frame(cancel_window, bg="#1f2937", height=50)
        header_frame.pack(fill=tk.X)

        header_label = tk.Label(header_frame, text="Cancel Booking", font=("Century Gothic", 16, "bold"),
                                fg="white", bg="#1f2937", pady=10)
        header_label.pack()

        # === Outer Content Frame ===
        outer_frame = tk.Frame(cancel_window, bg="#ecf0f1", padx=15, pady=20)
        outer_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(10, 20))

        # === Data Entry Section ===
        form_frame = tk.Frame(outer_frame, bg="#ecf0f1")
        form_frame.pack(fill=tk.X, pady=(0, 20))

        fields = [
            ("Booking ID", tk.Entry),
            ("Reason (Optional)", tk.Entry),
        ]
        self.entries = {}

        for i, (label_text, field_type) in enumerate(fields):
            label = tk.Label(form_frame, text=label_text, font=("Helvetica", 10, "bold"),
                            bg="#ecf0f1", fg="#2c3e50")
            label.grid(row=i, column=0, sticky="w", pady=5, padx=(5, 10))

            # Frame around each Entry
            entry_outer_frame = tk.Frame(form_frame, bg="white", relief="groove", borderwidth=2)
            entry_outer_frame.grid(row=i, column=1, sticky="ew", pady=5)

            entry = field_type(entry_outer_frame, font=("Helvetica", 10), width=28,
                            bg="white", fg="black", relief="flat")
            entry.pack(fill=tk.X, padx=4, pady=3)  # Smaller vertical padding!

            self.entries[label_text] = entry

        form_frame.columnconfigure(1, weight=1)

        # === Buttons Frame ===
        btn_frame = tk.Frame(outer_frame, bg="#ecf0f1")
        btn_frame.pack()

        submit_btn = RoundedButton(
            btn_frame,
            text="Submit",
            command=lambda: self.submit_cancel_booking(cancel_window),
            radius=14,
            padding=8,  # Slightly smaller buttons too
            color="#1abc9c",
            hover_color="#16a085",
            text_color="white",
            font=("Helvetica", 10, "bold"),
            border_color="#16a085",
            border_width=2,
        )
        submit_btn.grid(row=0, column=0, padx=15)

        cancel_btn = RoundedButton(
            btn_frame,
            text="Cancel",
            command=cancel_window.destroy,
            radius=14,
            padding=8,
            color="#e74c3c",
            hover_color="#c0392b",
            text_color="white",
            font=("Helvetica", 10, "bold"),
            border_color="#c0392b",
            border_width=2,
        )
        cancel_btn.grid(row=0, column=1, padx=15)





    def submit_cancel_booking(self, cancel_window):
        """Handles the booking cancellation request."""
        try:
            booking_id = self.entries["Booking ID"].get().strip()

            cancellation_reason = self.entries["Reason (Optional)"].get().strip()

            if not booking_id:
                cancel_window.grab_release()  # Release before showing message
                CTkMessagebox(title="Error", message="Please enter a Booking ID.", icon="cancel")
                return

            # API request
            api_url = f"http://127.0.0.1:8000/bookings/cancel/{booking_id}/"
            if cancellation_reason:
                api_url += f"?cancellation_reason={requests.utils.quote(cancellation_reason)}"

            headers = {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"}
            response = requests.post(api_url, headers=headers)

            # Handle response
            if response.status_code == 200:
                cancel_window.grab_release()  # Ensure grab is released
                canceled_booking = response.json().get("canceled_booking", {})

                # Success message
                msgbox = CTkMessagebox(
                    title="Success",
                    message=f"Booking {canceled_booking.get('id', booking_id)} canceled successfully!\n"
                            f"Room Status: {canceled_booking.get('room_status', 'N/A')}\n"
                            f"Booking Status: {canceled_booking.get('status', 'N/A')}\n"
                            f"Reason: {canceled_booking.get('cancellation_reason', 'None')}",
                    icon="check",
                    option_1="OK",
                )
                # Wait for user to press OK before closing the window
                if msgbox.get() == "OK":
                    cancel_window.destroy()


            else:
                cancel_window.grab_release()  # Release before showing message
                CTkMessagebox(title="Error", message=response.json().get("detail", "Cancellation failed."), icon="warning")

        except requests.exceptions.RequestException as e:
            cancel_window.grab_release()
            CTkMessagebox(title="Error", message=f"Request failed: {e}", icon="cancel")




    
    
    def clear_right_frame(self):
        for widget in self.right_frame.winfo_children():
            widget.pack_forget()

if __name__ == "__main__":
    root = tk.Tk()
    BookingManagement(root, token="dummy_token")
    root.mainloop()
    
    