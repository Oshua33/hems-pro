import os
import requests
import pandas as pd
import subprocess
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

import os
import sys


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)





TOKEN_FILE = "token.txt"
API_BASE_URL = resource_path("http://127.0.0.1:8000")  # Update this if needed
BASE_URL = f"{API_BASE_URL}/bookings"  # For booking-related endpoints



def save_token(token):
    """Save the token to a file."""
    with open(TOKEN_FILE, "w") as file:
        file.write(token)

def load_token():
    """Load the token from the file."""
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "r") as file:
            return file.read().strip()
    return None


def get_user_role(token):
    url = "http://127.0.0.1:8000/users/me"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)

    try:
        user_data = response.json()
        print("User Data Response:", user_data)  # Debugging line

        if not isinstance(user_data, dict):  # Ensure it's a dictionary
            print("Unexpected response format:", user_data)
            return "guest"

        return user_data.get("role", "guest")  # Default to "guest"
    
    except Exception as e:
        print("Error fetching user role:", e)
        return "guest"  # Safe fallback

import requests

import requests
import logging

def api_request(endpoint, method="GET", data=None, token=None, params=None):
    url = f"http://127.0.0.1:8000{endpoint}"
    headers = {"Authorization": f"Bearer {token}"} if token else {}

    try:
        if method == "GET":
            response = requests.get(url, headers=headers, params=params)
        elif method == "POST":
            response = requests.post(url, json=data, headers=headers)
        elif method == "PUT":
            response = requests.put(url, json=data, headers=headers)
        elif method == "DELETE":
            response = requests.delete(url, headers=headers)
        else:
            logging.error(f"Unsupported method: {method}")
            return None

        if response.status_code in [200, 201]:
            return response.json()
        elif response.status_code == 204:
            return {"detail": "No content"}
        else:
            try:
                error_data = response.json()
                logging.warning(f"Error: {response.status_code}, Response: {error_data}")
            except ValueError:
                logging.warning(f"Error: {response.status_code}, Response: {response.text}")
            return None

    except requests.exceptions.RequestException as e:
        logging.error(f"Request error: {e}")
        return None


import requests

def perform_booking_action(endpoint, data, token):
    """ Perform API requests for booking management """
    url = f"http://127.0.0.1:8000/bookings/{endpoint}"  # Adjust base URL if needed
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    
    try:
        response = requests.post(url, json=data, headers=headers)  # Change method if necessary
        return response.json()
    except Exception as e:
        return {"error": str(e)}
    


def export_to_excel(data, filename="payments_report.xlsx"):
    """Export data to a well-formatted Excel file."""
    
    if not data:
        return None  # No data to export

    try:
        df = pd.DataFrame(data)

        # Ensure filename has .xlsx extension
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        # Write to Excel with formatting
        df.to_excel(filename, index=False, sheet_name="Payments")

        # Load the workbook and sheet to apply styling
        wb = load_workbook(filename)
        ws = wb["Payments"]

        # Format header row: Bold and Centered
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column width based on content
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get the column letter (A, B, C...)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        # Save the formatted file
        wb.save(filename)

        return filename

    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return None


def print_excel(file_path):
    """Open and print an Excel file."""
    if not os.path.exists(file_path):
        messagebox.showerror("Error", "File not found!")
        return

    try:
        os.startfile(file_path, "print")  # Windows
        # subprocess.run(["lp", file_path])  # Mac/Linux alternative
        messagebox.showinfo("Printing", "Report is being printed.")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to print: {str(e)}")
    
   